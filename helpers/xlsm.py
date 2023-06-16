import pandas as pd
from openpyxl.formula.translate import Translator


def populate_sheet_series(df, sheet, col=2, row=2):
    df = df.fillna("")
    output = []
    start_row = row
    start_col = col

    if isinstance(df, pd.DataFrame):
        for index, row in df.iterrows():
            output.append(row.values.tolist())
        for rowy, row in enumerate(output, start=start_row):
            for colx, value in enumerate(row, start=start_col):
                sheet.cell(column=colx, row=rowy, value=value)
    else:
        output = df
        for row, value in enumerate(output, row):
            sheet.cell(row, col, value)


def populate_sap_data_sheet(df, sheet, start_col=2, start_row=2):
    df = df.fillna("")
    output = []

    for index, row in df.iterrows():
        output.append(row.values.tolist())

    for rowy, row in enumerate(output, start=start_row):
        for colx, value in enumerate(row, start=start_col):
            sheet.cell(column=colx, row=rowy, value=value)


def extend_concats(sheet, start_row=100, col_letter="A"):
    last_row = sheet.max_row
    i = start_row
    while i < last_row:
        i += 1
        formula = sheet[f"{col_letter}2"].value
        sheet[f"{col_letter}{i}"] = Translator(
            formula, origin=f"{col_letter}2"
        ).translate_formula(f"{col_letter}{i}")


def extend_values(sheet, start_row=100, col_letter="A"):
    last_row = sheet.max_row
    last_row_value = sheet[f"{col_letter}{last_row}"].value
    print(last_row_value)
    i = start_row
    while i < last_row:
        i += 1
        sheet[f"{col_letter}{i}"].value = last_row_value


def get_last_row(worksheet, col):
    for row in worksheet.iter_rows():
        if row[0].value == None:
            empty_row = row[0].row
