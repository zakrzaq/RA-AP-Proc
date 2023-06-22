import os
import time
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

from utils.helpers import await_char, use_dotenv
import utils.prompts as pr

from state.output import output

use_dotenv()


def load_log():
    try:
        start = time.time()
        log_file = os.environ["AP_LOG"]
        output.add(f"{pr.info}LOG loading...")
        wb = load_workbook(filename=log_file, keep_vba=True)
        end = time.time()
        output.add(f"{pr.done}LOG successfully loaded: {round(end - start, 2)}s")
        return wb
    except:
        output.add(f"{pr.cncl}LOG count not be loaded")


def save_log(log, server=True, name="TEST_output"):
    print(server)

    def save():
        start = time.time()
        output.add(f"{pr.file}LOG is being saved")
        log.save(os.environ["AP_LOG"])
        end = time.time()
        output.add(f"{pr.done}LOG successfully saved: {round(end - start, 2)}s")

    def test_save():
        start = time.time()
        output.add(f"{pr.file}TEST LOG is being saved")
        log.save(os.path.join(os.environ["DIR_OUT"], f"{name}.xlsm"))
        end = time.time()
        output.add(f"{pr.done}TEST LOG successfully saved: {round(end - start, 2)}s")

    if server:
        save()
    else:
        await_char(
            "y", f"{pr.prmt}Save actual LOG file? Press Y to continue.", test_save()
        )


def populate_sheet_series(df, sheet, col=2, row=2):
    df = df.fillna("")
    output = []
    start_row = row
    start_col = col

    if isinstance(df, pd.DataFrame):
        for index, row in df.iterrows():
            output.append(list(row.values))
        for rowy, row in enumerate(output, start=start_row):
            for colx, value in enumerate(row, start=start_col):
                sheet.cell(column=colx, row=rowy, value=value)
    else:
        output = df
        for row, value in enumerate(output, row):
            sheet.cell(row, col, value)


def populate_sap_data_sheet(df, sheet, start_col=2, start_row=2):
    df = df.fillna("")
    output = []

    for index, row in df.iterrows():
        output.append(row.values.tolist())

    for rowy, row in enumerate(output, start=start_row):
        for colx, value in enumerate(row, start=start_col):
            sheet.cell(column=colx, row=rowy, value=value)


def extend_concats(sheet, start_row=100, col_letter="A"):
    last_row = sheet.max_row
    i = start_row
    while i < last_row:
        i += 1
        formula = sheet[f"{col_letter}2"].value
        sheet[f"{col_letter}{i}"] = Translator(
            formula, origin=f"{col_letter}2"
        ).translate_formula(f"{col_letter}{i}")


def extend_values(sheet, start_row=100, col_letter="A"):
    last_row = sheet.max_row
    last_row_value = sheet[f"{col_letter}{last_row}"].value
    print(last_row_value)
    i = start_row
    while i < last_row:
        i += 1
        sheet[f"{col_letter}{i}"].value = last_row_value


def get_first_empty_row(worksheet: Worksheet, col: str):
    match col.lower():
        case "a":
            i = 0
        case "b":
            i = 1
        case _:
            i = 2

    if worksheet.max_row == 1048576:
        empty_row: int = 0
        for row in worksheet.iter_rows():
            if row[i].value == None:
                empty_row: int = int(row[i].row)
                return empty_row
        return empty_row
    else:
        worksheet.max_row
