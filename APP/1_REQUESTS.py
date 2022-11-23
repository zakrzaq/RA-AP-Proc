###

# LIBS

from openpyxl.formula.translate import Translator
from openpyxl import load_workbook
import pandas as pd
import warnings
import os
import keyboard
import dotenv
dotenv_file = dotenv.find_dotenv()
dotenv.load_dotenv(dotenv_file)

warnings.filterwarnings("ignore")

# VARIABLES
ready_to_save = False

# OPEN LOG FILE NAD GENERATE SHEETS VARIABLES

# ACTUAL
# log_file = 'W:\\AP MM Service Request Log.xlsm'
log_file = os.environ['AP_LOG']

# TESTING
# file = 'W:\\AP MM Service Request Log_CLEANOUT.xlsm'
# file = 'W:\\AP MM Service Request Log_NEW.xlsm'
# file = "W:\\VBA.xlsm"

log = load_workbook(filename=log_file, keep_vba=True)

ws_active = log['Active Materials']
ws_mara = log['mara']
ws_marc = log['marc']
ws_mvke = log['mvke']
ws_ausp = log['ausp']
ws_mlan = log['mlan']
ws_price = log['ZZ_MATPRC_HIST']
ws_gts = log['gts']
ws_text = log['sales text']

# ws_active["A1"].value

# IMPORT REQUESTS FORM DESKTOP
directory = os.environ["HOME_DIR"]

# print(directory)
requests = pd.DataFrame()

for filename in os.listdir(directory):
    if filename.endswith(".xlsm"):
        if 'AP_Material_Master_Service_Request_Form' in filename:
            file = os.path.join(directory, filename)
            print(filename)
            df = pd.read_excel(file, 2)
            df = df.iloc[1:, :]
            # print(df)
            requests = pd.concat([requests, df])

# cleanup data
if not requests.empty:
    requests['Unnamed: 3'] = requests['Unnamed: 3'].str.strip()
    requests['Unnamed: 2'] = requests['Unnamed: 2'].str.replace(
        '@rockwellautomation.com', '')
    requests['Unnamed: 2'] = requests['Unnamed: 2'].str.replace(
        '@ra.rockwell.com', '')
    requests['Unnamed: 2'] = requests['Unnamed: 2'].str.lower()
    requests.insert(0, 'Date', pd.to_datetime('today').strftime("%m/%d/%Y"))

    # POPULATE REQUEST TO LOG
    try:
        # last populated row
        for cell in ws_active['B']:
            # print(cell.value)
            if cell.value is None:
                # print(cell.row)
                ws_active_firstrow = cell.row
                break

        requests = requests.fillna('')
        requests_output = []
        for index, row in requests.iterrows():
            # print(row.values.tolist())
            requests_output.append(row.values.tolist())
        # print(requests_output)

        for rowy, row in enumerate(requests_output, start=ws_active_firstrow):
            for colx, value in enumerate(row, start=1):
                ws_active.cell(column=colx, row=rowy, value=value)

        for cell in ws_active['B']:
            if cell.value is None:
                ws_active_lastrow = cell.row
                break
    except:
        print('could not populate data ')

       # EXTEND FORMULAS By Col / Rows
    try:
        column_list = ['O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'Y', 'X', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH',
                       'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AY', 'AZ', 'BA', 'BB', 'BC', 'BD', 'BE']
        ws_active_lastrow -= 1
        for x in column_list:
            i = ws_active_firstrow - 1
            # print(ws_active_firstrow, ws_active_lastrow, x, i)
            while i < ws_active_lastrow:
                i += 1
                formula = ws_active['{}2'.format(x)].value
                # print(formula)
                ws_active['{}{}'.format(x, i)] = Translator(
                    formula, origin="{}2".format(x)).translate_formula("{}{}".format(x, i))
                # print("{}{}".format(x, i), ws_active['{}{}'.format(x, i)].value)
    except:
        print('could not extend formulas')

        # CREATE SORT
    try:
        sort_count = ws_active_lastrow + 1
        # print(sort_count)
        i = 2
        while i < sort_count:
            if i == 1:
                continue
            # print(i)
            ws_active["BF{}".format(i)].value = i - 1
            i += 1
    except:
        print('could not create sort order')

    ready_to_save = True

else:
    print('no request files found')

if ready_to_save:
    # TESTING
    log.save(os.path.join(os.environ['APP_DIR'],
             'OUTPUTS', 'test_requests.xlsm'))
    print('mock saved')
    # ACTUAL
    print('Press y tp save to actual AP LOG:')
    while True:
        if keyboard.is_pressed("y"):  # returns True if "y" is pressed
            print('saving...')
            log.save(os.environ['AP_LOG'])
            print('LOG saved')
            break  # break the while loop is "y" is pressed

# MAKE LIST OF MATERIALS IN AP LOG

active_matnr_list_file = os.path.join(
    os.environ['HOME_DIR'], 'AP materials.txt')

if os.path.exists(active_matnr_list_file):
    os.remove(active_matnr_list_file)
    print('material list redone')
else:
    print('material list not found')

active_matnr_list = ''
with open(active_matnr_list_file, "w") as file:
    for row in ws_active:
        if row[4].value != None:
            if 'SAP MATNR' in row[4].value:
                continue
            elif row[4].value.isnumeric():
                active_matnr_list += ('000000000000000000' +
                                      row[4].value + '\n')
            else:
                active_matnr_list += (row[4].value + '\n')
    # print(active_matnr_list)
    file.write(active_matnr_list)
    file.close()
