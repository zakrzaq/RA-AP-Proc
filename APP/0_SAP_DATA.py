# LIBS
from helpers import populate_sap_data_sheet, extend_concats
from openpyxl.formula.translate import Translator
from openpyxl import load_workbook
import pandas as pd
import warnings
import os
import dotenv
dotenv_file = dotenv.find_dotenv()
dotenv.load_dotenv(dotenv_file)

warnings.filterwarnings("ignore")

# VARIABLES

# OPEN LOG FILE NAD GENERATE SHEETS VARIABLES

# ACTUAL
log_file = os.environ['AP_LOG']
# TESTING
# log_file = 'W:\\AP MM Service Request Log_CLEANOUT.xlsm'
# log_file = 'W:\\AP MM Service Request Log_TESTING.xlsm'
# log_file = "W:\\VBA.xlsm"

log = load_workbook(filename=log_file, keep_vba=True)

ws_active = log['Active Materials']
ws_mara = log['mara']
ws_marc = log['marc']
ws_mvke = log['mvke']
ws_ausp = log['ausp']
ws_mlan = log['mlan']
ws_price = log['ZZ_MATPRC_HIST']
ws_gts = log['gts']
ws_text = log['sales text']
ws_list = [ws_mara, ws_marc, ws_mvke, ws_ausp,
           ws_mlan, ws_price, ws_gts, ws_text]

# # clean out data
for sheet in ws_list:
    sheet.delete_rows(100, ws_mara.max_row)

# DEFINE DATA FILES
mara = marc = mvke = ausp = mlan = price = gts = text = pd.DataFrame()

fl_mara = os.path.join(os.environ['APP_DIR'], 'DATA', 'mara.XLSX')
fl_marc = os.path.join(os.environ['APP_DIR'], 'DATA', 'marc.XLSX')
fl_mvke = os.path.join(os.environ['APP_DIR'], 'DATA', 'mvke.XLSX')
fl_ausp = os.path.join(os.environ['APP_DIR'], 'DATA', 'ausp.XLSX')
fl_mlan = os.path.join(os.environ['APP_DIR'], 'DATA', 'mlan.XLSX')
fl_price = os.path.join(os.environ['APP_DIR'], 'DATA', 'price.XLSX')
fl_gts = os.path.join(os.environ['APP_DIR'], 'DATA', 'gts.XLSX')
fl_text = os.path.join(os.environ['APP_DIR'], 'DATA', 'sales_text.XLSX')

# load sap data to df
if os.path.exists(fl_mara):
    mara = pd.read_excel(fl_mara)
if os.path.exists(fl_marc):
    marc = pd.read_excel(fl_marc)
if os.path.exists(fl_mvke):
    mvke = pd.read_excel(fl_mvke)
if os.path.exists(fl_ausp):
    ausp = pd.read_excel(fl_ausp)
if os.path.exists(fl_mlan):
    mlan = pd.read_excel(fl_mlan)
if os.path.exists(fl_price):
    price = pd.read_excel(fl_price)
if os.path.exists(fl_gts):
    gts = pd.read_excel(fl_gts)
if os.path.exists(fl_text):
    text = pd.read_excel(fl_text)

inputs_list = [mara, marc, mvke, ausp, mlan, price, gts, text]


# FUNCTIONS


inputs_not_empty = 0
for input in inputs_list:
    if not input.empty:
        inputs_not_empty += 1

if inputs_not_empty == 7:
    populate_sap_data_sheet(mara, ws_mara)
    populate_sap_data_sheet(marc, ws_marc)
    populate_sap_data_sheet(mvke, ws_mvke)
    populate_sap_data_sheet(ausp, ws_ausp)
    populate_sap_data_sheet(mlan, ws_mlan)
    populate_sap_data_sheet(price, ws_price)
    populate_sap_data_sheet(gts, ws_gts)
    populate_sap_data_sheet(text, ws_text)

    for sheet in ws_list:
        extend_concats(sheet)

    # save test
    log.save(os.path.join(os.environ['TMP_OUT_DIR'], 'test_sap_data.xlsm'))
    # log.save(os.environ['AP_LOG'])

else:
    print('Missing SAP data')
